class War:
    def __init__(self,number_of_cards,player_pack,number_of_players):
        self.number_of_cards=number_of_cards
        self.player_pack=player_pack                
        self.number_of_players=number_of_players 
    def shuffle(self):
        number_of_cards=self.number_of_cards
        pack=[]
        for card in range(number_of_cards):
            for card_suit in range(4):
                    pack.append(card+1)  
        self.pack_after_shuffle=pack
        return(pack)   
    def deal(self):
        pack=self.pack_after_shuffle
        index=self.player_pack
        import random
        random.shuffle(pack)
        player_1=pack[index:]
        player_2=pack[:index]
        self.player_pack_1=player_1
        self.player_pack_2=player_2
        pass
    def game_two_players(self):
        import random
        player_1=self.player_pack_1
        player_2=self.player_pack_2
        new_player_1=[]
        new_player_2=[]
        while True:
            if (len(player_1)!=0 and len(player_2)!=0 and (len(new_player_1)==0 or len(new_player_2)==0)):
                for card_player_1, card_player_2 in zip(player_1, player_2):
                    if card_player_1>card_player_2:
                            player_1.remove(card_player_1)
                            player_2.remove(card_player_2)
                            new_player_1.append(card_player_1)
                            new_player_1.append(card_player_2)
                    elif  card_player_1<card_player_2:
                            player_1.remove(card_player_1)
                            player_2.remove(card_player_2)
                            new_player_2.append(card_player_1)
                            new_player_2.append(card_player_2)
                    else:
                            player_1.remove(card_player_1)
                            player_2.remove(card_player_2)
                    random.shuffle(new_player_1)
                    random.shuffle(new_player_2)
            elif (len(player_1)!=0) and (len(new_player_2)!=0):
                for card_player_1, card_player_2 in zip( player_1, new_player_2):
                    if card_player_1>card_player_2:
                            player_1.remove(card_player_1)
                            new_player_2.remove(card_player_2)
                            new_player_1.append(card_player_1)
                            new_player_1.append(card_player_2)
                    elif  card_player_1<card_player_2:
                            player_1.remove(card_player_1)
                            new_player_2.remove(card_player_2)
                            player_2.append(card_player_1)
                            player_2.append(card_player_2)
                    else:
                            player_1.remove(card_player_1)
                            new_player_2.remove(card_player_2)
                    random.shuffle(player_1)
                    random.shuffle(new_player_2)           
            elif (len(player_2)!=0) and (len(new_player_1)!=0):
                for card_player_1, card_player_2 in zip(new_player_1,player_2):
                    if card_player_1>card_player_2:
                            new_player_1.remove(card_player_1)
                            player_2.remove(card_player_2)
                            player_1.append(card_player_1)
                            player_1.append(card_player_2)
                    elif  card_player_1<card_player_2:
                            new_player_1.remove(card_player_1)
                            player_2.remove(card_player_2)
                            new_player_2.append(card_player_1)
                            new_player_2.append(card_player_2)
                    else:
                            new_player_1.remove(card_player_1)
                            player_2.remove(card_player_2)
                    random.shuffle(new_player_1)
                    random.shuffle(player_2)            
            elif (len(player_1)==0) and (len(player_2)==0) and ((len(new_player_1)!=0) and (len(new_player_2)!=0)):   
                for card_player_1, card_player_2 in zip( new_player_1, new_player_2):
                    if card_player_1>card_player_2:
                            new_player_1.remove(card_player_1)
                            new_player_2.remove(card_player_2)
                            player_1.append(card_player_1)
                            player_1.append(card_player_2)
                    elif  card_player_1<card_player_2:
                            new_player_1.remove(card_player_1)
                            new_player_2.remove(card_player_2)
                            player_2.append(card_player_1)
                            player_2.append(card_player_2)
                    else:
                            new_player_1.remove(card_player_1)
                            new_player_2.remove(card_player_2)
                    random.shuffle(new_player_1)
                    random.shuffle(new_player_2)
            else:
                break
        self.end_player_1=player_1
        self.end_player_2=player_2
        self.end_player_new_pack_1=new_player_1
        self.end_player_new_pack_2=new_player_2
    def display(self):
        if(self.number_of_players==2):
            player_1=self.end_player_1
            player_2=self.end_player_2
            new_player_1=self.end_player_new_pack_1
            new_player_2=self.end_player_new_pack_2
            print("Number of cards for player 1: ",(len(player_1)+len(new_player_1)))
            print("Number of cards for player 2: ",(len(player_2)+len(new_player_2)))
            if len(player_1)!=0 and len(new_player_1)!=0 and len(player_2)==0 and len(new_player_2)==0:
                print("Player 1 won")
            elif len(player_1)==0 and len(new_player_1)==0 and len(player_2)!=0 and len(new_player_2)!=0:
                print("Player 2 won")
        elif(self.number_of_players==3):
            player_1=self.end_player_1
            player_2=self.end_player_2
            new_player_1=self.end_player_new_pack_1
            new_player_2=self.end_player_new_pack_2
            print("Number of cards for player 1: ",(len(player_1)+len(new_player_1)))
            print("Number of cards for player 2: ",(len(player_2)+len(new_player_2)))
            if len(player_1)!=0 and len(new_player_1)!=0 and len(player_2)==0 and len(new_player_2)==0:
                print("Player 1 won")
            elif len(player_1)==0 and len(new_player_1)==0 and len(player_2)!=0 and len(new_player_2)!=0:
                print("Player 2 won")
        elif(self.number_of_players==4):
            player_1=self.end_player_1
            player_2=self.end_player_2
            new_player_1=self.end_player_new_pack_1
            new_player_2=self.end_player_new_pack_2
            print("Number of cards for player 1: ",(len(player_1)+len(new_player_1)))
            print("Number of cards for player 2: ",(len(player_2)+len(new_player_2)))
            if len(player_1)!=0 and len(new_player_1)!=0 and len(player_2)==0 and len(new_player_2)==0:
                print("Player 1 won")
            elif len(player_1)==0 and len(new_player_1)==0 and len(player_2)!=0 and len(new_player_2)!=0:
                print("Player 2 won")
        elif(self.number_of_players==5):
            player_1=self.end_player_1
            player_2=self.end_player_2
            new_player_1=self.end_player_new_pack_1
            new_player_2=self.end_player_new_pack_2
            print("Number of cards for player 1: ",(len(player_1)+len(new_player_1)))
            print("Number of cards for player 2: ",(len(player_2)+len(new_player_2)))
            if len(player_1)!=0 and len(new_player_1)!=0 and len(player_2)==0 and len(new_player_2)==0:
                print("Player 1 won")
            elif len(player_1)==0 and len(new_player_1)==0 and len(player_2)!=0 and len(new_player_2)!=0:
                print("Player 2 won")
        elif(self.number_of_players==6):
            player_1=self.end_player_1
            player_2=self.end_player_2
            new_player_1=self.end_player_new_pack_1
            new_player_2=self.end_player_new_pack_2
            print("Number of cards for player 1: ",(len(player_1)+len(new_player_1)))
            print("Number of cards for player 2: ",(len(player_2)+len(new_player_2)))
            if len(player_1)!=0 and len(new_player_1)!=0 and len(player_2)==0 and len(new_player_2)==0:
                print("Player 1 won")
            elif len(player_1)==0 and len(new_player_1)==0 and len(player_2)!=0 and len(new_player_2)!=0:
                print("Player 2 won")
    def excel(self):
        import openpyxl
        workbook = openpyxl.load_workbook('table.xlsx')
        sheet = workbook.active
        score_1=len(self.end_player_1+self.end_player_new_pack_1)
        score_2=len(self.end_player_2+self.end_player_new_pack_2)
        score_3=len(self.end_player_3+self.end_player_new_pack_3)
        score_4=len(self.end_player_4+self.end_player_new_pack_4)
        score_5=len(self.end_player_5+self.end_player_new_pack_5)
        score_6=len(self.end_player_6+self.end_player_new_pack_6)
        sheet["A1"]="Player 1 score" 
        sheet["B1"]="Player 2 score"
        sheet["C1"]="Player 3 score"
        sheet["D1"]="Player 4 score"
        sheet["E1"]="Player 5 score"
        sheet["F1"]="Player 6 score"
        data=[score_1,score_2,score_3,score_4,score_5,score_6]
        sheet.append(data)
        workbook.save('table.xlsx')
    def clear(self):
        import openpyxl
        workbook = openpyxl.load_workbook('table.xlsx')
        sheet = workbook.active
        for i in range(self.number_of_players):
            column_index = i
            for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
                for cell in row:
                    cell.value = None
        workbook.save('table.xlsx')
             
class Condition:
     def __init__(self,number_of_players):
          self.number_of_players=number_of_players
     def check(self):
          if(self.number_of_players==2):
               return self.check_value
          elif(self.number_of_players==3):
               pass
          elif(self.number_of_players==4):
               pass
          elif(self.number_of_players==5):
               pass
          elif(self.number_of_players==6):
               pass








